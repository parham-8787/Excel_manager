from openpyxl import Workbook, load_workbook

class Excel:
    def __init__(self, address_file, sheet_name, *, create_workbook="no"):
        self.address_file = (fr"{address_file}.xlsx" if not address_file.endswith(".xlsx") else fr"{address_file}")
        if create_workbook == "yes":
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = sheet_name
            self.save()
            print("New workbook created.")
        elif create_workbook == "no":
            try:
                self.wb = load_workbook(self.address_file)
                self.ws = self.wb[sheet_name]
                print("Workbook loaded successfully.")
            except FileNotFoundError:
                print("File not found, creating a new workbook...")
                self.wb = Workbook()
                self.ws = self.wb.active
                self.ws.title = sheet_name
                self.save()
            except KeyError:
                print(f"Sheet '{sheet_name}' not found, using active sheet.")
                self.ws = self.wb.active
            except Exception as e:
                print("Error loading workbook:", e)
            else:
                self.save()
        else:
            raise ValueError("create_workbook must be 'yes' or 'no'")

    def save(self):
        """Save workbook with correct extension"""
        self.wb.save(self.address_file)

    def create_sheet(self, sheet_name):
        try:
            self.wb.create_sheet(title=sheet_name)
        except Exception as e:
            print("Error creating sheet:", e)
        else:
            self.save()
            print("Sheet created...")

    def rename_sheet(self, sheet_name, new_sheet_name):
        try:
            sheet = self.wb[sheet_name]
            sheet.title = new_sheet_name
        except Exception as e:
            print("Error renaming sheet:", e)
        else:
            self.save()
            print("Sheet renamed...")

    def remove_sheet(self, sheet_name):
        try:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)
        except Exception as e:
            print("Error removing sheet:", e)
        else:
            self.save()
            print("Sheet removed...")

    def show_sheets(self):
        try:
            sheets_name = self.wb.sheetnames
        except Exception as e:
            print("Error fetching sheets:", e)
        else:
            print(" | ".join(sheets_name))

    def insert_data(self, data):
        try:
            if not data:
                print("No data to insert.")
                return
            if isinstance(data[0], (list, tuple)):
                for row in data:
                    self.ws.append(row)
            else:
                self.ws.append(data)
        except Exception as e:
            print("Error inserting data:", e)
        else:
            self.save()
            print("Data appended...")

    def show_data(self, min_row, max_row, min_column, max_column):
        try:
            data_show = []
            for row in self.ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_column,
                max_col=max_column,
                values_only=True,
            ):
                data_show.append(row)
                print(*row)
        except Exception as e:
            print("Error showing data:", e)
        else:
            return data_show

    def excel_to_excel(self, min_row, max_row, min_col, max_col, target_file):
        target_file = (
            f"{target_file}.xlsx"
            if not target_file.endswith(".xlsx")
            else target_file
        )
        try:
            wb2 = load_workbook(target_file)
            ws2 = wb2.active
        except FileNotFoundError:
            print("Target file not found, creating a new workbook...")
            wb2 = Workbook()
            ws2 = wb2.active
        try:
            for row in self.ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                ws2.append(row)
            wb2.save(target_file)
        except Exception as e:
            print("Error copying data:", e)
        else:
            print("Data copied successfully to target file.")
